"""
Excel数据导出模块
将统计数据写入Excel文件
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Dict, List
import os


class ExcelWriter:
    def __init__(self, filename: str = 'danmaku_statistics.xlsx'):
        self.filename = filename
        
    def write_statistics(self, stats: Dict):
        """
        将统计数据写入Excel
        """
        # 创建DataFrame
        data = {
            '排名': [],
            '弹幕内容': [],
            '出现次数': []
        }
        
        for item in stats['top_8_danmaku']:
            data['排名'].append(item['rank'])
            data['弹幕内容'].append(item['danmaku'])
            data['出现次数'].append(item['count'])
        
        df = pd.DataFrame(data)
        
        # 写入Excel
        with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='弹幕统计', index=False)
            
            # 获取工作表并格式化
            worksheet = writer.sheets['弹幕统计']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 60
            worksheet.column_dimensions['C'].width = 15
            
            # 设置标题样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置数据对齐
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                row[0].alignment = Alignment(horizontal='center')  # 排名
                row[1].alignment = Alignment(horizontal='left', wrap_text=True)  # 弹幕内容
                row[2].alignment = Alignment(horizontal='center')  # 次数
            
            # 添加汇总信息
            summary_row = worksheet.max_row + 2
            worksheet[f'A{summary_row}'] = '总计'
            worksheet[f'B{summary_row}'] = f"有效弹幕总数: {stats['total_count']}"
            worksheet[f'C{summary_row}'] = f"原始弹幕数: {stats['original_count']}"
            
            worksheet[f'A{summary_row}'].font = Font(bold=True)
            worksheet[f'B{summary_row}'].font = Font(bold=True)
            worksheet[f'C{summary_row}'].font = Font(bold=True)
        
        print(f"统计数据已保存到: {self.filename}")


if __name__ == '__main__':
    writer = ExcelWriter()
    test_stats = {
        'total_count': 100,
        'original_count': 150,
        'top_8_danmaku': [
            {'rank': 1, 'danmaku': '大模型真厉害', 'count': 20},
            {'rank': 2, 'danmaku': 'GPT很好用', 'count': 15}
        ]
    }
    writer.write_statistics(test_stats)

